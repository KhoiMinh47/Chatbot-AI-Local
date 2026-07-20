"""Parser adapters with stable errors and normalized provenance metadata."""

from __future__ import annotations

import csv
import hashlib
import io
import logging
import os
import re
import tempfile
import threading
import unicodedata
from abc import ABC, abstractmethod
from functools import lru_cache
from importlib.metadata import version
from pathlib import Path
from typing import Any, ClassVar, Protocol
from uuid import UUID

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from worker.domain import ElementType, NormalizedDocument, NormalizedElement

_DOCLING_CONVERT_LOCK = threading.Lock()


class ParserError(ValueError):
    """A stable parser failure safe to expose through ingestion status."""

    def __init__(self, code: str, message: str, *, retryable: bool = False) -> None:
        super().__init__(message)
        self.code = code
        self.retryable = retryable


class ParserAdapter(Protocol):
    """Protocol for document parsers."""

    def supports_mime_type(self, mime_type: str) -> bool: ...

    async def parse(
        self,
        content: bytes,
        filename: str,
        mime_type: str,
        document_id: UUID,
        version_id: UUID,
        tenant_id: UUID,
    ) -> NormalizedDocument: ...


class BaseParser(ABC):
    """Base parser with conservative text and language helpers."""

    @abstractmethod
    def supports_mime_type(self, mime_type: str) -> bool:
        """Return whether this adapter accepts a MIME type."""

    @abstractmethod
    async def parse(
        self,
        content: bytes,
        filename: str,
        mime_type: str,
        document_id: UUID,
        version_id: UUID,
        tenant_id: UUID,
    ) -> NormalizedDocument:
        """Parse content into the normalized domain schema."""

    @staticmethod
    def _clean_text(text: str) -> str:
        normalized = unicodedata.normalize("NFC", text).replace("\r\n", "\n").replace("\r", "\n")
        return "\n".join(line.rstrip() for line in normalized.split("\n")).strip()

    @staticmethod
    def _detect_language(text: str) -> str | None:
        try:
            import langdetect

            sample = text[:2000].strip()
            return langdetect.detect(sample) if sample else None
        except Exception:
            return None


def _label_value(item: Any) -> str:
    label = getattr(item, "label", "text")
    return str(getattr(label, "value", label)).lower()


def _element_type(label: str) -> ElementType:
    if label in {"title", "section_header", "field_heading"} or "heading" in label:
        return ElementType.HEADING
    if label == "table":
        return ElementType.TABLE
    if label in {"list_item", "checkbox_selected", "checkbox_unselected"}:
        return ElementType.LIST
    if label in {"caption", "footnote"}:
        return ElementType.CAPTION
    if label in {"code", "formula"}:
        return ElementType.CODE
    return ElementType.PARAGRAPH


def _confidence_from_metadata(value: object) -> float | None:
    if isinstance(value, dict):
        for key, nested in value.items():
            if "confidence" in str(key).lower() and isinstance(nested, int | float):
                confidence = float(nested)
                if 0 <= confidence <= 1:
                    return confidence
            found = _confidence_from_metadata(nested)
            if found is not None:
                return found
    elif isinstance(value, list):
        for nested in value:
            found = _confidence_from_metadata(nested)
            if found is not None:
                return found
    return None


def _docling_provenance(item: Any) -> tuple[int | None, list[float] | None, dict[str, Any]]:
    raw_provenance = getattr(item, "prov", None)
    provenance = raw_provenance if isinstance(raw_provenance, list) else []
    page: int | None = None
    bbox: list[float] | None = None
    if provenance:
        first = provenance[0]
        raw_page = getattr(first, "page_no", None)
        if isinstance(raw_page, int) and raw_page > 0:
            page = raw_page
        raw_bbox = getattr(first, "bbox", None)
        coordinates = [getattr(raw_bbox, name, None) for name in ("l", "t", "r", "b")]
        if all(isinstance(value, int | float) for value in coordinates):
            bbox = [float(value) for value in coordinates if isinstance(value, int | float)]

    metadata: dict[str, Any] = {
        "docling_label": _label_value(item),
        "provenance_count": len(provenance),
    }
    raw_meta = getattr(item, "meta", None)
    if raw_meta is not None:
        dumped = raw_meta.model_dump(mode="json") if hasattr(raw_meta, "model_dump") else raw_meta
        confidence = _confidence_from_metadata(dumped)
        if confidence is not None:
            metadata["ocr_confidence"] = confidence
    return page, bbox, metadata


@lru_cache(maxsize=2)
def _document_converter(ocr_enabled: bool = False) -> Any:
    """Build one converter per worker process with explicit local RapidOCR."""

    from docling.datamodel.base_models import InputFormat
    from docling.datamodel.pipeline_options import PdfPipelineOptions, RapidOcrOptions
    from docling.document_converter import DocumentConverter, PdfFormatOption

    pdf_options = PdfPipelineOptions(
        artifacts_path=os.environ.get("DOCLING_ARTIFACTS_PATH"),
        # Native text is the fast first pass. A second cached converter is used
        # only when page coverage proves that OCR is necessary.
        do_ocr=ocr_enabled,
        ocr_options=RapidOcrOptions(),
        enable_remote_services=False,
        allow_external_plugins=False,
    )
    return DocumentConverter(
        format_options={InputFormat.PDF: PdfFormatOption(pipeline_options=pdf_options)}
    )


def _docling_page_count(document: Any) -> int | None:
    pages = getattr(document, "pages", None)
    if isinstance(pages, dict | list | tuple) and pages:
        return len(pages)
    return None


def _docling_elements(
    document: Any,
    document_id: UUID,
    version_id: UUID,
    mime_type: str,
) -> list[NormalizedElement]:
    elements: list[NormalizedElement] = []
    headings: dict[int, str] = {}
    pptx_mime = "application/vnd.openxmlformats-officedocument.presentationml.presentation"
    for item, level in document.iterate_items():
        label = _label_value(item)
        element_type = _element_type(label)
        if element_type == ElementType.TABLE and hasattr(item, "export_to_markdown"):
            raw_text = item.export_to_markdown(doc=document)
        else:
            raw_text = getattr(item, "text", "")
        text = BaseParser._clean_text(str(raw_text))
        if not text:
            continue

        heading_level = max(int(level), 0)
        if element_type == ElementType.HEADING:
            headings = {depth: value for depth, value in headings.items() if depth < heading_level}
            headings[heading_level] = text
        section_path = [headings[depth] for depth in sorted(headings)]

        page, bbox, metadata = _docling_provenance(item)
        slide = page if mime_type == pptx_mime else None
        page_number = None if mime_type == pptx_mime else page
        metadata["reading_level"] = heading_level
        elements.append(
            NormalizedElement(
                element_id=f"{document_id}:{version_id}:docling:{page or 0}:{len(elements)}",
                type=element_type,
                text=text,
                page=page_number,
                slide=slide,
                section_path=section_path,
                bbox=bbox,
                metadata=metadata,
            )
        )
    return elements


class DoclingParser(BaseParser):
    """Docling adapter for PDF, DOCX and PPTX, including tables and provenance."""

    _SUPPORTED_MIMES: ClassVar[set[str]] = {
        "application/pdf",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        "text/html",
    }
    _PPTX_MIME = "application/vnd.openxmlformats-officedocument.presentationml.presentation"

    def supports_mime_type(self, mime_type: str) -> bool:
        return mime_type in self._SUPPORTED_MIMES

    async def parse(
        self,
        content: bytes,
        filename: str,
        mime_type: str,
        document_id: UUID,
        version_id: UUID,
        tenant_id: UUID,
    ) -> NormalizedDocument:
        if not content:
            raise ParserError("PARSER_EMPTY_FILE", "The uploaded document is empty.")

        suffix = Path(filename).suffix.lower()
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as temporary:
            temporary.write(content)
            temporary_path = Path(temporary.name)

        try:
            try:
                # The converter owns heavyweight OCR/layout models. Reuse it in
                # each Celery process and serialize access if a threaded pool is
                # selected accidentally.
                with _DOCLING_CONVERT_LOCK:
                    result = _document_converter().convert(temporary_path)
            except Exception as parse_exc:
                # Log the actual error for debugging
                logger = logging.getLogger(__name__)
                logger.error(f"Docling parsing failed for {filename}: {parse_exc}", exc_info=True)
                raise ParserError(
                    "PARSER_CORRUPT_DOCUMENT",
                    f"The document could not be parsed: {str(parse_exc)[:200]}",
                ) from parse_exc

            document = result.document
            expected_units = _docling_page_count(document)
            elements = _docling_elements(document, document_id, version_id, mime_type)
            covered_units = {
                location
                for element in elements
                if (location := (element.slide if mime_type == self._PPTX_MIME else element.page))
                is not None
            }
            coverage = (
                len(covered_units) / expected_units
                if expected_units is not None and expected_units > 0
                else 1.0
                if elements
                else 0.0
            )
            ocr_used = False
            ocr_attempted = False
            ocr_warning: str | None = None
            if mime_type == "application/pdf" and coverage < 0.90:
                ocr_attempted = True
                try:
                    with _DOCLING_CONVERT_LOCK:
                        ocr_result = _document_converter(True).convert(temporary_path)
                    ocr_document = ocr_result.document
                    ocr_elements = _docling_elements(
                        ocr_document,
                        document_id,
                        version_id,
                        mime_type,
                    )
                    ocr_covered = {
                        element.page for element in ocr_elements if element.page is not None
                    }
                    if len(ocr_covered) > len(covered_units) or sum(
                        len(element.text) for element in ocr_elements
                    ) > sum(len(element.text) for element in elements):
                        document = ocr_document
                        elements = ocr_elements
                        covered_units = ocr_covered
                        expected_units = _docling_page_count(document) or expected_units
                        ocr_used = True
                except Exception:
                    logging.getLogger(__name__).exception("Selective PDF OCR retry failed")
                    ocr_warning = "selective_ocr_failed"

            if not elements:
                raise ParserError(
                    "PARSER_EMPTY_OUTPUT",
                    "No readable text was found; use a clearer scan or verify OCR support.",
                )
            language = self._detect_language(" ".join(element.text for element in elements[:20]))
            return NormalizedDocument(
                document_id=document_id,
                version_id=version_id,
                tenant_id=tenant_id,
                source_name=filename,
                mime_type=mime_type,
                language=language,
                content_hash=hashlib.sha256(content).hexdigest(),
                elements=elements,
                metadata={
                    "parser": "docling",
                    "parser_version": version("docling"),
                    "ocr_engine": "rapidocr-onnxruntime",
                    "ocr_attempted": ocr_attempted,
                    "ocr_enabled": ocr_used,
                    "ocr_warning": ocr_warning,
                    "force_full_page_ocr": False,
                    "remote_services_enabled": False,
                    "expected_unit_count": expected_units,
                    "covered_unit_count": len(covered_units),
                },
            )
        finally:
            temporary_path.unlink(missing_ok=True)


class CSVParser(BaseParser):
    """CSV adapter that repeats headers and retains exact row ranges."""

    def supports_mime_type(self, mime_type: str) -> bool:
        return mime_type in {"text/csv", "application/csv"}

    async def parse(
        self,
        content: bytes,
        filename: str,
        mime_type: str,
        document_id: UUID,
        version_id: UUID,
        tenant_id: UUID,
    ) -> NormalizedDocument:
        try:
            decoded = content.decode("utf-8-sig")
            rows = list(csv.reader(io.StringIO(decoded), strict=True))
        except (UnicodeDecodeError, csv.Error):
            raise ParserError(
                "PARSER_INVALID_CSV",
                "The CSV is not valid UTF-8 or contains malformed quoting.",
            ) from None
        if not rows or not rows[0]:
            raise ParserError("PARSER_EMPTY_FILE", "The CSV has no header row.")

        header = rows[0]
        data_rows = rows[1:]
        elements: list[NormalizedElement] = []
        row_batch_size = 20
        for offset in range(0, max(len(data_rows), 1), row_batch_size):
            batch = data_rows[offset : offset + row_batch_size]
            lines = [
                "| " + " | ".join(header) + " |",
                "|" + "|".join("---" for _ in header) + "|",
                *("| " + " | ".join(row) + " |" for row in batch),
            ]
            row_start = offset + 1
            row_end = offset + len(batch)
            elements.append(
                NormalizedElement(
                    element_id=f"{document_id}:{version_id}:csv:{row_start}:{row_end}",
                    type=ElementType.TABLE,
                    text="\n".join(lines),
                    sheet=filename,
                    cell_range=f"rows:{row_start}-{row_end}",
                    section_path=[f"Rows {row_start}-{row_end}"],
                    metadata={"row_start": row_start, "row_end": row_end},
                )
            )
        sample = " ".join(element.text[:500] for element in elements[:3])
        return NormalizedDocument(
            document_id=document_id,
            version_id=version_id,
            tenant_id=tenant_id,
            source_name=filename,
            mime_type=mime_type,
            language=self._detect_language(sample),
            content_hash=hashlib.sha256(content).hexdigest(),
            elements=elements,
            metadata={"parser": "ntc-csv", "parser_version": "1"},
        )


class XLSXParser(BaseParser):
    """Memory-bounded worksheet parser retaining sheet and cell-range provenance."""

    _MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def supports_mime_type(self, mime_type: str) -> bool:
        return mime_type == self._MIME

    async def parse(
        self,
        content: bytes,
        filename: str,
        mime_type: str,
        document_id: UUID,
        version_id: UUID,
        tenant_id: UUID,
    ) -> NormalizedDocument:
        if not content:
            raise ParserError("PARSER_EMPTY_FILE", "The spreadsheet is empty.")
        try:
            workbook = load_workbook(
                io.BytesIO(content),
                read_only=True,
                data_only=False,
            )
        except Exception as error:
            raise ParserError(
                "PARSER_INVALID_XLSX",
                f"The spreadsheet could not be parsed: {str(error)[:160]}",
            ) from error

        elements: list[NormalizedElement] = []
        covered_sheets: set[str] = set()
        batch_size = 25
        try:
            expected_sheet_count = len(workbook.sheetnames)
            for worksheet in workbook.worksheets:
                element_count_before_sheet = len(elements)
                rows = worksheet.iter_rows(values_only=True)
                try:
                    first = next(rows)
                except StopIteration:
                    continue
                width = max(1, len(first))
                headers = tuple(
                    str(value) if value is not None else f"Column {index}"
                    for index, value in enumerate(first, 1)
                )
                batch: list[tuple[object, ...]] = []
                row_start = 2

                def append_batch(
                    values: list[tuple[object, ...]],
                    start: int,
                    *,
                    bound_headers: tuple[str, ...] = headers,
                    bound_width: int = width,
                    sheet_name: str = worksheet.title,
                ) -> None:
                    if not values:
                        return
                    end = start + len(values) - 1
                    rendered_rows = [
                        ["" if value is None else str(value) for value in row] for row in values
                    ]
                    if not any(any(cell for cell in row) for row in rendered_rows):
                        return
                    lines = [
                        "| " + " | ".join(bound_headers) + " |",
                        "|" + "|".join("---" for _ in bound_headers) + "|",
                        *("| " + " | ".join(row) + " |" for row in rendered_rows),
                    ]
                    cell_range = f"A{start}:{get_column_letter(bound_width)}{end}"
                    elements.append(
                        NormalizedElement(
                            element_id=(
                                f"{document_id}:{version_id}:xlsx:{sheet_name}:{cell_range}"
                            ),
                            type=ElementType.TABLE,
                            text="\n".join(lines),
                            sheet=sheet_name,
                            cell_range=cell_range,
                            section_path=[sheet_name],
                            metadata={"row_start": start, "row_end": end},
                        )
                    )
                    covered_sheets.add(sheet_name)

                for row_number, row in enumerate(rows, start=2):
                    batch.append(tuple(row))
                    if len(batch) >= batch_size:
                        append_batch(batch, row_start)
                        batch = []
                        row_start = row_number + 1
                append_batch(batch, row_start)
                if len(elements) == element_count_before_sheet and any(
                    value is not None and str(value).strip() for value in first
                ):
                    cell_range = f"A1:{get_column_letter(width)}1"
                    elements.append(
                        NormalizedElement(
                            element_id=(
                                f"{document_id}:{version_id}:xlsx:{worksheet.title}:{cell_range}"
                            ),
                            type=ElementType.TABLE,
                            text="| " + " | ".join(headers) + " |",
                            sheet=worksheet.title,
                            cell_range=cell_range,
                            section_path=[worksheet.title],
                            metadata={"row_start": 1, "row_end": 1},
                        )
                    )
                    covered_sheets.add(worksheet.title)
        finally:
            workbook.close()

        if not elements:
            raise ParserError("PARSER_EMPTY_OUTPUT", "No readable spreadsheet cells were found.")
        sample = " ".join(element.text[:500] for element in elements[:3])
        return NormalizedDocument(
            document_id=document_id,
            version_id=version_id,
            tenant_id=tenant_id,
            source_name=filename,
            mime_type=mime_type,
            language=self._detect_language(sample),
            content_hash=hashlib.sha256(content).hexdigest(),
            elements=elements,
            metadata={
                "parser": "ntc-openpyxl-read-only",
                "parser_version": version("openpyxl"),
                "expected_unit_count": expected_sheet_count,
                "covered_unit_count": len(covered_sheets),
            },
        )


class SourceCodeParser(BaseParser):
    """UTF-8 source parser using stable, overlapping line windows and symbol hints."""

    _MIMES: ClassVar[set[str]] = {
        "application/json",
        "text/javascript",
        "text/typescript",
        "text/x-c",
        "text/x-c++src",
        "text/x-go",
        "text/x-java-source",
        "text/x-python",
        "text/x-rust",
        "text/x-shellscript",
        "text/x-sql",
        "text/yaml",
    }
    _SYMBOL = re.compile(
        r"^\s*(?:async\s+def|def|class|function|func|fn|interface|type|CREATE\s+(?:TABLE|FUNCTION))\s+([A-Za-z_][\w$]*)",
        re.IGNORECASE,
    )

    def supports_mime_type(self, mime_type: str) -> bool:
        return mime_type in self._MIMES

    async def parse(
        self,
        content: bytes,
        filename: str,
        mime_type: str,
        document_id: UUID,
        version_id: UUID,
        tenant_id: UUID,
    ) -> NormalizedDocument:
        try:
            decoded = unicodedata.normalize("NFC", content.decode("utf-8-sig"))
        except UnicodeDecodeError:
            raise ParserError(
                "PARSER_INVALID_TEXT_ENCODING",
                "Source code must use UTF-8 encoding.",
            ) from None
        lines = decoded.replace("\r\n", "\n").replace("\r", "\n").splitlines()
        if not any(line.strip() for line in lines):
            raise ParserError("PARSER_EMPTY_FILE", "The source file is empty.")

        elements: list[NormalizedElement] = []
        window_size = 120
        overlap = 20
        start = 0
        while start < len(lines):
            end = min(len(lines), start + window_size)
            window = lines[start:end]
            symbol = next(
                (match.group(1) for line in window if (match := self._SYMBOL.match(line))),
                None,
            )
            text = "\n".join(window).strip("\n")
            if text.strip():
                line_start = start + 1
                line_end = end
                elements.append(
                    NormalizedElement(
                        element_id=(f"{document_id}:{version_id}:source:{line_start}:{line_end}"),
                        type=ElementType.CODE,
                        text=text,
                        line_start=line_start,
                        line_end=line_end,
                        section_path=[filename, *([symbol] if symbol else [])],
                        metadata={"symbol": symbol, "source_mime": mime_type},
                    )
                )
            if end >= len(lines):
                break
            start = end - overlap

        return NormalizedDocument(
            document_id=document_id,
            version_id=version_id,
            tenant_id=tenant_id,
            source_name=filename,
            mime_type=mime_type,
            language=self._detect_language(decoded),
            content_hash=hashlib.sha256(content).hexdigest(),
            elements=elements,
            metadata={
                "parser": "ntc-source-lines",
                "parser_version": "1",
                "expected_unit_count": 1,
                "covered_unit_count": 1,
            },
        )


class PlainTextParser(BaseParser):
    """Plain-text and Markdown adapter with heading breadcrumbs."""

    def supports_mime_type(self, mime_type: str) -> bool:
        return mime_type in {"text/plain", "text/markdown", "text/x-markdown"}

    async def parse(
        self,
        content: bytes,
        filename: str,
        mime_type: str,
        document_id: UUID,
        version_id: UUID,
        tenant_id: UUID,
    ) -> NormalizedDocument:
        try:
            decoded = unicodedata.normalize("NFC", content.decode("utf-8-sig"))
        except UnicodeDecodeError:
            raise ParserError(
                "PARSER_INVALID_TEXT_ENCODING",
                "The text document must use UTF-8 encoding.",
            ) from None
        paragraphs = [part.strip() for part in decoded.replace("\r\n", "\n").split("\n\n")]
        paragraphs = [part for part in paragraphs if part]
        if not paragraphs:
            raise ParserError("PARSER_EMPTY_FILE", "The text document is empty.")

        elements: list[NormalizedElement] = []
        headings: dict[int, str] = {}
        for index, paragraph in enumerate(paragraphs):
            first_line = paragraph.splitlines()[0].strip()
            marker_length = len(first_line) - len(first_line.lstrip("#"))
            is_heading = marker_length > 0 or (index == 0 and len(paragraph) < 100)
            element_type = ElementType.HEADING if is_heading else ElementType.PARAGRAPH
            text = self._clean_text(paragraph.lstrip("# ") if is_heading else paragraph)
            if is_heading:
                depth = marker_length or 1
                headings = {level: value for level, value in headings.items() if level < depth}
                headings[depth] = text
            section_path = [headings[level] for level in sorted(headings)]
            elements.append(
                NormalizedElement(
                    element_id=f"{document_id}:{version_id}:paragraph:{index}",
                    type=element_type,
                    text=text,
                    section_path=section_path,
                )
            )
        return NormalizedDocument(
            document_id=document_id,
            version_id=version_id,
            tenant_id=tenant_id,
            source_name=filename,
            mime_type=mime_type,
            language=self._detect_language(decoded),
            content_hash=hashlib.sha256(content).hexdigest(),
            elements=elements,
            metadata={"parser": "ntc-plain-text", "parser_version": "1"},
        )


class ParserRegistry:
    """Registry of available parser adapters."""

    def __init__(self) -> None:
        self._parsers: list[ParserAdapter] = [
            DoclingParser(),
            XLSXParser(),
            CSVParser(),
            SourceCodeParser(),
            PlainTextParser(),
        ]

    def get_parser(self, mime_type: str) -> ParserAdapter:
        for parser in self._parsers:
            if parser.supports_mime_type(mime_type):
                return parser
        raise ParserError(
            "PARSER_UNSUPPORTED_MIME",
            f"No parser available for MIME type {mime_type!r}.",
        )


parser_registry = ParserRegistry()

__all__ = [
    "CSVParser",
    "DoclingParser",
    "ParserAdapter",
    "ParserError",
    "ParserRegistry",
    "PlainTextParser",
    "SourceCodeParser",
    "XLSXParser",
    "parser_registry",
]
